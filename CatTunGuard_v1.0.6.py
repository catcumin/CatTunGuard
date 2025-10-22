# Copyright 2025 猫宁孜 (catcumin)
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import requests
import re
import time
import os
import maskpass
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class FRPViolationChecker:
    def __init__(self, base_api: str, auth_token: str):
        self.base_api = base_api
        self.headers = {
            "Authorization": auth_token,
            "Content-Type": "application/json"
        }
        self.config = {
            "timeout": 8,
            "max_workers": 5,
            "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
            "violation_keywords": ["色情", "赌博", "枪支", "违法", "私服", "外挂"],
            "web_local_ports": {"80", "8080", "8000", "443", "8888", "9000"},
            "html_indicators": {"<html", "<head", "<body", "<title", "<meta"}
        }
        self.excel_styles = {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_bg": "2F75B5",
            "violation_font": Font(color="FF0000"),
            "normal_font": Font(color="000000"),
            "border": Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
        }
        self.error_count = 0  # 检测流程错误计数器
        self.max_errors = 5   # 检测流程最大错误次数限制

    def fetch_all_tunnels(self) -> Optional[List[Dict]]:
        all_tunnels = []
        page = 1
        page_size = 10

        while True:
            api_url = f"{self.base_api}&page={page}&page_size={page_size}"
            try:
                resp = requests.get(api_url, headers=self.headers, timeout=15)
                resp.raise_for_status()
                data = resp.json()

                if data.get("code") != 200:
                    self.error_count += 1
                    print(f"❌ 数据获取失败：{data.get('msg', '未知错误')}（错误次数：{self.error_count}/{self.max_errors}）")
                    if self.error_count >= self.max_errors:
                        print("❌ 错误次数已达上限，即将退出程序")
                        return None
                    break

                tunnels = data.get("proxies", [])
                if not tunnels:
                    print(f"✅ 已获取所有隧道（共{len(all_tunnels)}个）")
                    break
                all_tunnels.extend(tunnels)
                print(f"已获取第{page}页隧道（累计{len(all_tunnels)}个）")

                pagination = data.get("pagination", {})
                total_pages = pagination.get("pages", 1)
                if page >= total_pages:
                    print(f"✅ 已获取所有{total_pages}页隧道（共{len(all_tunnels)}个）")
                    break

                page += 1
                time.sleep(1)

            except Exception as e:
                self.error_count += 1
                print(f"❌ 数据获取失败：{str(e)}（错误次数：{self.error_count}/{self.max_errors}）")
                if self.error_count >= self.max_errors:
                    print("❌ 错误次数已达上限，即将退出程序")
                    return None
                break

        return all_tunnels if all_tunnels else None

    def parse_link(self, link: str, proxy_type: str) -> Dict:
        link = link.strip()
        if link.startswith(("http://", "https://")):
            protocol = link.split("://")[0]
            address = link.split("://")[1]
        else:
            protocol = "https" if proxy_type == "https" else "http"
            address = link
        return {"protocol": protocol, "address": address, "full_url": f"{protocol}://{address}"}

    def check_web_content(self, full_url: str) -> Dict:
        result = {
            "is_web": False,
            "is_violation": False,
            "evidence": [],
            "error": None
        }

        try:
            resp = requests.get(
                url=full_url,
                headers={"User-Agent": self.config["user_agent"]},
                timeout=self.config["timeout"],
                allow_redirects=True,
                verify=False
            )
            content = resp.text.lower()

            if any(indicator in content for indicator in self.config["html_indicators"]):
                result["is_web"] = True
                result["evidence"].append(f"含网页特征（状态码：{resp.status_code}）")

            for kw in self.config["violation_keywords"]:
                if kw in content:
                    result["is_violation"] = True
                    result["evidence"].append(f"内容含违规关键词：{kw}")

        except requests.exceptions.RequestException as e:
            result["error"] = f"访问失败：{str(e)}"
        return result

    def analyze_tunnel(self, tunnel: Dict) -> Optional[Dict]:
        try:
            tunnel_id = tunnel["id"]
            username = tunnel["username"]
            proxy_type = tunnel["proxy_type"].lower()
            link = tunnel["link"]
            domain = tunnel["domain"].strip()
            local_port = str(tunnel["local_port"])

            need_check = False
            if proxy_type in ["http", "https"]:
                need_check = True
            elif proxy_type == "tcp" and local_port in self.config["web_local_ports"]:
                need_check = True
            elif domain:
                need_check = True

            if not need_check:
                return None

            link_info = self.parse_link(link, proxy_type)
            check_result = self.check_web_content(link_info["full_url"])

            domain_evidence = []
            if domain:
                if not re.match(r"^\d+\.\d+\.\d+\.\d+$", domain):
                    domain_evidence.append(f"绑定域名：{domain}（需确认备案状态）")
                else:
                    domain_evidence.append(f"绑定IP域名：{domain}")

            all_evidence = check_result["evidence"] + domain_evidence
            if check_result["error"]:
                all_evidence.append(f"检测异常：{check_result['error']}")

            is_violation = check_result["is_violation"] or (proxy_type in ["http", "https"] and domain and not re.match(r"^\d+\.\d+\.\d+\.\d+$", domain))

            return {
                "tunnel_id": tunnel_id,
                "username": username,
                "proxy_type": proxy_type,
                "link": link,
                "local_port": local_port,
                "domain": domain,
                "is_violation": is_violation,
                "evidence": "；".join(all_evidence),
                "check_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        except Exception as e:
            self.error_count += 1
            print(f"❌ 隧道分析失败：{str(e)}（错误次数：{self.error_count}/{self.max_errors}）")
            if self.error_count >= self.max_errors:
                print("❌ 错误次数已达上限，即将退出程序")
            return None

    def export_to_excel(self, results: List[Dict]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "FRP隧道检测结果"

        headers = [
            "序号", "隧道ID", "用户名", "代理类型", "外网地址(link)",
            "内网端口", "绑定域名", "是否违规", "违规证据", "检测时间"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.excel_styles["header_font"]
            cell.fill = PatternFill(start_color=self.excel_styles["header_bg"], end_color=self.excel_styles["header_bg"], fill_type="solid")
            cell.border = self.excel_styles["border"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=row-1).border = self.excel_styles["border"]
            ws.cell(row=row, column=2, value=result["tunnel_id"]).border = self.excel_styles["border"]
            ws.cell(row=row, column=3, value=result["username"]).border = self.excel_styles["border"]
            ws.cell(row=row, column=4, value=result["proxy_type"]).border = self.excel_styles["border"]
            ws.cell(row=row, column=5, value=result["link"]).border = self.excel_styles["border"]
            ws.cell(row=row, column=6, value=result["local_port"]).border = self.excel_styles["border"]
            ws.cell(row=row, column=7, value=result["domain"] or "无").border = self.excel_styles["border"]
            violation_cell = ws.cell(
                row=row,
                column=8,
                value="是" if result["is_violation"] else "否"
            )
            violation_cell.font = self.excel_styles["violation_font"] if result["is_violation"] else self.excel_styles["normal_font"]
            violation_cell.border = self.excel_styles["border"]
            violation_cell.alignment = Alignment(horizontal="center")
            evidence_cell = ws.cell(row=row, column=9, value=result["evidence"])
            evidence_cell.border = self.excel_styles["border"]
            evidence_cell.alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=10, value=result["check_time"]).border = self.excel_styles["border"]

        column_widths = [5, 10, 15, 10, 30, 10, 25, 8, 60, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"frp_violation_check_{timestamp}.xlsx"
        wb.save(filename)

        return filename

    def run_detection(self):
        print("====== 开始FRP隧道违规检测 ======")
        
        all_tunnels = self.fetch_all_tunnels()
        if not all_tunnels or self.error_count >= self.max_errors:
            print("====== 检测中断（无有效隧道数据或错误次数超限） ======")
            return

        results = []
        with ThreadPoolExecutor(max_workers=self.config["max_workers"]) as executor:
            future_to_tunnel = {
                executor.submit(self.analyze_tunnel, tunnel): tunnel
                for tunnel in all_tunnels
            }

            total = len(all_tunnels)
            for i, future in enumerate(as_completed(future_to_tunnel), 1):
                if self.error_count >= self.max_errors:
                    print("\n❌ 错误次数已达上限，终止检测")
                    break
                tunnel = future_to_tunnel[future]
                try:
                    res = future.result(timeout=self.config["timeout"] + 2)
                    if res:
                        results.append(res)
                except Exception as e:
                    self.error_count += 1
                    print(f"\n❌ 隧道ID {tunnel['id']} 处理失败：{str(e)}（错误次数：{self.error_count}/{self.max_errors}）")
                    if self.error_count >= self.max_errors:
                        print("❌ 错误次数已达上限，即将退出程序")
                        break
                print(f"处理进度：{i}/{total}（隧道ID: {tunnel['id']}）", end="\r")

        if self.error_count >= self.max_errors:
            return

        if results:
            excel_file = self.export_to_excel(results)
            print(f"\n\n✅ 检测结果已导出至：{os.path.abspath(excel_file)}")
        else:
            print("\n\nℹ️  无需要记录的检测结果（未发现需关注的隧道）")

        print("\n====== 检测结果汇总 ======")
        violation_count = sum(1 for r in results if r["is_violation"])
        print(f"总检测隧道数：{len(results)} | 违规隧道数：{violation_count} | 需关注隧道数：{len(results) - violation_count}")
        print("====== 检测结束 ======")

    def verify_token(self) -> bool:
        try:
            verify_url = f"{self.base_api}&page=1&page_size=1"
            resp = requests.get(verify_url, headers=self.headers, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            return data.get("code") == 200
        except Exception:
            return False


if __name__ == "__main__":
    os.environ["PYTHONUTF8"] = "1"
    
    line = "=" * 40
    print(line)
    project_name = "喵盾 FRP 隧检器 | CatTunGuard"
    print(f"       {project_name}       ")
    print(line)
    print("  开发者：catcumin (猫宁孜)")
    print("  版本：v1.0.6")
    print("  功能：FRP隧道违规使用检测工具")
    print("  更新日期：2025-10-22")
    print("  邮箱：969008120@qq.com")
    print("  GitHub：https://github.com/catcumin")
    print(line + "\n")

    BASE_API = "https://console.frp.api.xhuzim.top/api/v1/admin/proxies?status=online"
    AUTH_TOKEN = ""
    token_attempts = 0  # Token输入错误次数（从0开始计数）
    max_token_attempts = 5  # Token错误上限为5次

    # Token验证循环
    while token_attempts < max_token_attempts:
        if not AUTH_TOKEN:
            print("请输入管理员认证token（用于API访问）：")
        else:
            remaining_attempts = max_token_attempts - token_attempts
            print(f"\n请重新输入有效的管理员认证token（剩余尝试次数：{remaining_attempts}）：")
        
        AUTH_TOKEN = maskpass.askpass(prompt="> ", mask="*").strip()
        temp_checker = FRPViolationChecker(BASE_API, AUTH_TOKEN)
        
        if temp_checker.verify_token():
            print("✅ Token验证通过，开始检测流程...\n")
            break
        
        token_attempts += 1
        print(f"❌ 无效的Token（错误次数：{token_attempts}/{max_token_attempts}）")

    # 若错误次数达上限，提示退出
    if token_attempts >= max_token_attempts:
        print("❌ Token尝试次数已达上限，即将退出程序")
        input("\n按Enter键关闭窗口...")
        exit(1)

    # 执行检测流程（仅添加KeyboardInterrupt捕获，其他逻辑不变）
    try:
        checker = FRPViolationChecker(base_api=BASE_API, auth_token=AUTH_TOKEN)
        checker.run_detection()
    except KeyboardInterrupt:
        # 专门处理手动中断（如Ctrl+C）
        print("\n\n⚠️  程序被手动中断")
    except Exception as e:
        print(f"\n⚠️  程序运行异常：{str(e)}")
    finally:
        if 'checker' in locals() and checker.error_count < checker.max_errors:
            input("\n检测已结束，按Enter键关闭窗口...")
        else:
            input("\n按Enter键关闭窗口...")